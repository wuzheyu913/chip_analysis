import tejapi
from datetime import date
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

import pygsheets

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
from email.mime.image import MIMEImage
from pathlib import Path
from email.mime.application import MIMEApplication 
# ==================================================================
# API 拿
# 今天的日期
today = date.today().strftime("%Y-%m-%d")
# today = '2024-04-18'

tejapi.ApiConfig.api_key = "eJ8bhkYk3BEbRVtqNxRc0R45ZzCYaM"
tejapi.ApiConfig.api_base = "http://10.10.10.66"

# 需要的欄位
columns = ['coid', 'mdate', 'key3', 'zvol_d', 'zamt_d', 'net_s1', 'net_s1b', 'net_a1', 'net_a1b']
# columns = ['key3', 'zvol_d', 'zamt_d', 'net_s1', 'net_s1b']
# data = tejapi.get('TWN/ABSR20', mdate='2024-01-09', opts={'columns': columns}, chinese_column_name=True, paginate=True)
data = tejapi.get('TWN/ABSR20', mdate=today, opts={'columns': columns}, chinese_column_name=True, paginate=True)

# ==================================================================
# 排除分點清單
# 排除 (外資) (總公司)
lst_bname_foreign = ['摩根大通', '上海匯豐' ,'美林', '台灣摩根', '台灣匯立', '法銀巴黎', '花旗環球', '美林', '美商高盛', '港商野村', '港麥格理', '瑞銀', '聯邦']
lst_bname_corporate = ['元大', '兆豐', '元富', '台新', '永豐金', '玉山', '合庫', '亞東', '康和', '統一', '凱基', '凱基台北', '富邦', '富邦經紀', '群益','華南永昌'
                      ,'中國信託', '臺銀', '宏遠', '彰銀', '第一金證', '新光', '土銀', '國泰綜合', '大和國泰', '元大經紀', '企銀']
lst_exclude = lst_bname_foreign + lst_bname_corporate

# ==================================================================
# main fuc

# -----------------------------------------
# 1. 整理
def arrange_data(df):
    df['年月日'] = df['年月日'].dt.strftime('%Y-%m-%d')
    df = df[df['證券名稱'].apply(lambda x: len(str(x)) == 4)]
    df.columns = ['scode', 'date', '排名', 'vol(千股)', 'vol(萬)', 'buy_量', '券商分點_量', 'buy_金額', '券商分點_金額']

    # 把 公司中文名 merge
    df_name = pd.read_excel('comp_name.xlsx')
    df_name.rename(columns={'代號': 'scode', '名稱':'sname'}, inplace=True)
    df_name['scode'] = df_name['scode'].astype(str)
    df = df.merge(df_name, on=['scode'], how='left')
    df = df[df['sname'].notnull()]
    df = df[['scode', 'sname', 'date', '排名', 'vol(千股)', 'vol(萬)', 'buy_量', '券商分點_量', 'buy_金額', '券商分點_金額']]

    # 整理買賣超金額 / 股數 / 買超佔比
    df['buy_proportion'] = round(df['buy_量'] / df['vol(千股)']*100,2)
    df['vol(千股)'] = round(df['vol(千股)']/1000, 2)
    df['vol(萬)'] = round(df['vol(萬)']/10000, 2)
    df['buy_量'] = round(df['buy_量']/1000, 2)
    df['buy_金額'] = round(df['buy_金額']/10000, 2)

    # 整理 成交量
    df_temp = df[['scode', 'date', 'buy_量', '券商分點_量']].copy()
    df = df[['scode', 'sname', 'date', '排名', 'vol(千股)', 'vol(萬)', 'buy_金額', '券商分點_金額', 'buy_proportion']]
    df = df.merge(df_temp, left_on=['scode', 'date', '券商分點_金額'], right_on=['scode', 'date', '券商分點_量'], how='left')
    df = df[['scode', 'sname', 'date', '排名', 'vol(千股)', 'vol(萬)', 'buy_量', 'buy_金額', '券商分點_金額', 'buy_proportion']]
    df.rename(columns={'券商分點_金額': 'bname'}, inplace=True)

    # 篩選
    df = df[df['buy_金額']>100]

    df.reset_index(drop=True, inplace=True)
    return df

df = data.copy()
df = arrange_data(df)

# 條件 
# 1 : 買超金額 > ?
# 2 : 買超佔比 > ?
# 3 : 排除分點
df_out = df[(df['buy_金額']>300)&(~df['bname'].isin(lst_exclude))].copy()
df_out = df_out.sort_values(['buy_金額'],ascending=False).reset_index(drop=True)
df_out['scode'] = df_out['scode'].astype(int)

# -----------------------------------------
# 1.1 標出可能是地緣的
def mark_same_locate(df_out):
    # 分點地址
    gc = pygsheets.authorize(service_file='sotck-chip-analysis-4d71166d853b.json')
    sheet = gc.open('chip')
    # print(sheet.worksheets()) # 看有哪些 sheet
    df_chip_info = pd.DataFrame(sheet[4].get_all_records())
    df_chip_info['代號'] = df_chip_info['代號'].astype(str)
    df_chip_info = df_chip_info[df_chip_info['TEJ_證券商名稱']!=''].reset_index(drop=True)
    df_chip_info['分點地址'] = df_chip_info['分點地址'].apply(lambda x:x[:3])

    # 公司地址
    comp_loc = pd.read_excel('公司所在縣市.xlsx')

    # 合併
    df_chip_info['分點地址'] = df_chip_info['分點地址'].apply(lambda x: x[:2])
    df_out = df_out.merge(df_chip_info[['TEJ_證券商名稱', '分點地址']], left_on='bname', right_on='TEJ_證券商名稱', how='left')
    df_out = df_out.merge(comp_loc, on='scode', how='left')

    # 標出有可能是地緣的
    df_out.loc[(df_out['分點地址']==df_out['公司縣市'])&(df_out['分點地址']!='台北'), '可能是地緣'] = 'Y'
    df_out.drop(columns=['TEJ_證券商名稱', '分點地址', '公司縣市'], inplace=True)
    return df_out

df_out = mark_same_locate(df_out)

# -----------------------------------------
# 2 標特別分點
def get_googlesheet_chip():
    gc = pygsheets.authorize(service_file='sotck-chip-analysis-4d71166d853b.json')
    sheet = gc.open('chip')
    print(sheet.worksheets()) # 看有哪些 sheet
    df_chip_google = pd.DataFrame(sheet[0].get_all_records())
    df_chip_chairman = pd.DataFrame(sheet[1].get_all_records())
    return df_chip_google, df_chip_chairman

def mark_special_chip(df, df_chip_google, df_chip_chairman):
    # TEJ 的分點都是4個字，所以要調整 ==
    dic = {'第一金': '第一', '華南永昌': '華南', '敦北法人': '敦北', '群益金鼎':'群益', '永豐金':'永豐', '中國信託':'中信', '台灣企銀':'台企', 
          '北高雄':'北高', '忠孝鼎富':'忠鼎', '西台中':'西台', '彰化民生':'彰生', '大天母':'大天', '高美館':'高美', '南京東路':'南京', '蘆洲中正':'蘆洲', '統一新台中':'統一新台',
          '北成功':'北成', '北三重':'北重', '大稻埕':'稻埕',' ':''}

    # 特別分點
    df_c = df_chip_google.copy()
    df_c['分點'] = df_c['分點'].replace(dic, regex=True)
    df_c = df_c.rename(columns={'分點':'bname'})
    df_c = df_c[['scode', 'bname', '特性', 'Good', '權證', '年度1', '年度2', '備註']]
    df['特別記錄'] = ''
    df = df.merge(df_c, on=['scode', 'bname'], how='left')
    df.loc[df['特性'].notnull(),'特別記錄'] = 'Y'
    df = df.sort_values(by=['特性', 'buy_金額'], ascending=[True, False])

    # 董監分點
    dfchairman = df_chip_chairman.copy()
    dfchairman = dfchairman.rename(columns={'分點':'bname'})
    dfchairman['bname'] = dfchairman['bname'].replace(dic, regex=True)
    dfchairman = dfchairman[['scode', '申報人身分', '姓名', 'bname', '年度', '厲害']]
    dfchairman['申報人身分'] = dfchairman['申報人身分'].str.replace('\t','')
    df['董監分點'] = ''
    df = df.merge(dfchairman, on=['scode','bname'], how='left')
    df.loc[df['申報人身分'].notnull(),'董監分點'] = 'Y'
    df = df.sort_values(by=['特別記錄', '董監分點', 'buy_金額'], ascending=[False, False, False])
    df = df.reset_index(drop=True)
    return df

df_chip_google, df_chip_chairman = get_googlesheet_chip()
df_out = mark_special_chip(df_out, df_chip_google, df_chip_chairman)

# -----------------------------------------
# to excel
file_name = '分點買賣超_%s.xlsx'%(date.today().strftime("%Y%m%d"))
df_out.to_excel(file_name, index=False, header=True)

# -----------------------------------------
# 3. 把 excel 著色
def color_excel(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.worksheets[0]

    # 顯示 cell 資料
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            cell = sheet.cell(row=i, column=j)
            if i == 1:
                cell.font = Font(name='微軟正黑體', bold=True, size=12)
            elif j in [1,3,4,5,6,7,11,12,14,15,16,17,19,22,23]:
                cell.font = Font(name='Calibri')
            elif j in [2,9,13,18,20,21]:
                cell.font = Font(name='微軟正黑體')
            if j in [12,13,14,15,16,17,18] and i == 1:
                color_fill = PatternFill(start_color='FFFF6F', fill_type='solid')
                sheet.cell(row=i, column=j).fill = color_fill
            elif j in [19,20,21,22,23] and i == 1:
                color_fill = PatternFill(start_color='FFAF60', fill_type='solid')
                sheet.cell(row=i, column=j).fill = color_fill
                
            if j == 8 and i != 1: # buy_金額
                # print(i, j, cell.value)
                if cell.value > 20000:
                    color_fill = PatternFill(start_color='ffd6cf', fill_type='solid')
                    sheet.cell(row=i, column=j).fill = color_fill          
                elif cell.value > 10000:
                    color_fill = PatternFill(start_color='ffe6b3', fill_type='solid')
                    sheet.cell(row=i, column=j).fill = color_fill
                elif cell.value > 5000:
                    color_fill = PatternFill(start_color='c7dbf7', fill_type='solid')
                    sheet.cell(row=i, column=j).fill = color_fill
                elif cell.value > 1000:
                    color_fill = PatternFill(start_color='DDF3FF', fill_type='solid')
                    sheet.cell(row=i, column=j).fill = color_fill
                cell.font = Font(name='Calibri')
            elif j == 10 and i != 1: # buy_proportion
                if cell.value > 50:
                    color_fill = PatternFill(start_color='bfa87c', fill_type='solid')
                    sheet.cell(row=i, column=j).fill = color_fill
                elif cell.value > 20:
                    color_fill = PatternFill(start_color='e6d395', fill_type='solid')
                    sheet.cell(row=i, column=j).fill = color_fill
                elif cell.value > 10:
                    color_fill = PatternFill(start_color='f0e6c0', fill_type='solid')
                    sheet.cell(row=i, column=j).fill = color_fill
                elif cell.value > 5:
                    color_fill = PatternFill(start_color='fffad9', fill_type='solid')
                    sheet.cell(row=i, column=j).fill = color_fill
                cell.font = Font(name='Calibri')
            elif j in [11,12,19] and i != 1: # 特別記錄分點
                if cell.value == 'Y':
                    color_fill = PatternFill(start_color='E6E6F2', fill_type='solid')
                    sheet.cell(row=i, column=j).fill = color_fill

    workbook.save(file_name)
    
color_excel(file_name)


# ==================================================================
# 自動寄信

def auto_send_mail(my_mail, receive_mail, file_name):
    content = MIMEMultipart()  #建立MIMEMultipart物件
    content["subject"] = "%s 籌碼分點excel"%(date.today().strftime("%Y-%m-%d"))   #郵件標題
    content["from"] = my_mail  #寄件者
    content["to"] = receive_mail #收件者
    content.attach(MIMEText("~~~今天ㄉ籌碼分點excel~~~"))  #郵件內容

    #寄送csv檔案
    
    with open(file_name, 'rb') as file:
        pdfload = MIMEApplication(file.read())
        pdfload.add_header('Content-Disposition', 'attachment', filename=file_name)
        content.attach(pdfload)

    with smtplib.SMTP(host="smtp.gmail.com", port="587") as smtp:  # 設定SMTP伺服器
        try:
            smtp.ehlo()  # 驗證SMTP伺服器
            smtp.starttls()  # 建立加密傳輸
            smtp.login(my_mail, "hwvo kclo yqfo lmvu")  # 登入寄件者gmail
            smtp.send_message(content)  # 寄送郵件
            print(receive_mail, "成功傳送")
        except Exception as e:
            print(receive_mail, "Error message: ", e)

my_mail = 'auto.momoju@gmail.com'
mail_list = pd.read_excel('mail_receive.xlsx')['收件人'].tolist()
for receive in mail_list:
    print(receive)
    auto_send_mail(my_mail, receive, file_name)
auto_send_mail(my_mail, receive_mail_2, file_name)

# ==================================================================